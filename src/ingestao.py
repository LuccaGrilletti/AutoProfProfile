"""Ingestão da planilha de entrada → dataclasses de atribuição.

Duas fontes de dados na v1:
  - Turmas: vêm de `carregar_planilha`. Cada professor (coluna `login`) é associado às turmas
    `turma`+`periodo` de suas linhas, COMBINADAS com o conjunto fixo de anos `config.ANOS_FIXOS`.
    As colunas de ano da planilha (`COLUNAS_ANOS`) NÃO são lidas na v1 — ver ADR-006.
  - Matérias: vêm de `carregar_materias` (aba `Materias` da planilha ou, na ausência dela, da lista
    fixa `MATERIAS_PADRAO`). NÃO são derivadas do Gestor de Classes no fluxo principal — ver ADR-006.

Linhas com o mesmo `login` são unidas (todas as turmas do professor entram na mesma atribuição).
"""

from dataclasses import dataclass, field
import src.config as config
import openpyxl

# Layout das colunas de ano da planilha → (ano, segmento). LEGADO: descreve o cabeçalho do arquivo,
# mas `carregar_planilha` não usa essas colunas na v1 (os anos vêm de config.ANOS_FIXOS). Ver ADR-006.
COLUNAS_ANOS = [
    ("1° EF", 1, "EF"), ("2° EF", 2, "EF"), ("3° EF", 3, "EF"),
    ("4° EF", 4, "EF"), ("5° EF", 5, "EF"), ("6° EF", 6, "EF"),
    ("7° EF", 7, "EF"), ("8° EF", 8, "EF"), ("9° EF", 9, "EF"),
    ("1° EM", 1, "EM"), ("2° EM", 2, "EM"), ("3° EM", 3, "EM"),
]


@dataclass(frozen=True)
class TurmaAlvo:
    ano: int          # 1..9 para EF, 1..3 para EM
    segmento: str     # "EF" | "EM"
    turma: str        # letra, ex: "G"
    periodo: str      # "Matutino" | "Vespertino" | ...

    @property
    def chave_pesquisa(self) -> str:
        """Termo usado no campo de busca do perfil do professor."""
        return f"Turma {self.turma} - {self.periodo}"

    @property
    def chave_canonica(self) -> str:
        """Chave legível para logs."""
        return f"{self.ano}° {self.segmento} - Turma {self.turma} - {self.periodo}"


@dataclass
class AtribuicaoProfessor:
    login: str
    turmas: list[TurmaAlvo] = field(default_factory=list)
    # As matérias NÃO entram aqui — são uma lista fixa global (ver carregar_materias), não por professor.


def carregar_planilha(caminho: str) -> list[AtribuicaoProfessor]:
    """Lê a aba `professores` e retorna uma AtribuicaoProfessor por login.

    Para cada linha (`login`, `turma`, `periodo`), gera uma TurmaAlvo para CADA ano de
    `config.ANOS_FIXOS` — ou seja, todo professor recebe o mesmo conjunto fixo de anos naquela
    turma/período (v1; as colunas de ano da planilha são ignoradas — ver ADR-006). Linhas sem
    `login`, `turma` ou `periodo` são ignoradas; linhas com o mesmo `login` têm suas turmas unidas.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb["professores"]
    agrupado: dict[str, list[TurmaAlvo]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        login, turma, periodo = row[0], row[1], row[2]
        if not login or not turma or not periodo:
            continue
        login, turma, periodo = str(login).strip(), str(turma).strip(), str(periodo).strip()
        turmas = agrupado.setdefault(login, [])
        for ano, segmento in config.ANOS_FIXOS:
            turmas.append(TurmaAlvo(ano=ano, segmento=segmento,
                                    turma=turma, periodo=periodo))
    wb.close()
    return [AtribuicaoProfessor(login=login, turmas=turmas)
            for login, turmas in agrupado.items()]


# Lista fixa de matérias (nome, série) usada quando a planilha não traz a aba `Materias`.
# É a fonte de matérias da v1 (ver ADR-006); para alterar o currículo, edite aqui ou a aba `Materias`.
MATERIAS_PADRAO: list[tuple[str, str]] = [
    ("Língua Portuguesa", "1º ano EF AI"),
    ("Matemática", "1º ano EF AI"),
    ("Inglês", "1º ano EF AI"),
    ("Bilingual Education", "1º ano EF AI"),
    ("Língua Portuguesa", "5º ano EF AI"),
    ("Matemática", "5º ano EF AI"),
    ("Ciências", "5º ano EF AI"),
    ("Geografia", "5º ano EF AI"),
    ("História", "5º ano EF AI"),
    ("Arte", "5º ano EF AI"),
    ("Bilingual Education", "5º ano EF AI"),
    ("Língua Portuguesa", "6º ano EF AF"),
    ("Matemática", "6º ano EF AF"),
    ("Ciências", "6º ano EF AF"),
    ("História", "6º ano EF AF"),
    ("Arte", "6º ano EF AF"),
    ("Inglês", "6º ano EF AF"),
    ("Bilingual Education", "6º ano EF AF"),
    ("Ciências", "7º ano EF AF"),
    ("Matemática", "1º ano EM"),
    ("Geografia", "1º ano EM"),
    ("História", "1º ano EM"),
    ("Sociologia", "1º ano EM"),
    ("Filosofia", "1º ano EM"),
    ("Física", "1º ano EM"),
    ("Química", "1º ano EM"),
    ("Biologia", "1º ano EM"),
    ("Arte", "1º ano EM"),
    ("Inglês", "1º ano EM"),
    ("Gramática", "1º ano EM"),
    ("Literatura", "1º ano EM"),
    ("Bilingual Education", "1º ano EM"),
    ("Matemática", "3º ano EM"),
    ("Ciências", "3º ano EM"),
    ("Geografia", "3º ano EM"),
    ("História", "3º ano EM"),
    ("Sociologia", "3º ano EM"),
    ("Filosofia", "3º ano EM"),
    ("Física", "3º ano EM"),
    ("Química", "3º ano EM"),
    ("Biologia", "3º ano EM"),
    ("Arte", "3º ano EM"),
    ("Inglês", "3º ano EM"),
    ("Gramática", "3º ano EM"),
    ("Literatura", "3º ano EM"),
    ("Bilingual Education", "3º ano EM"),
]

def carregar_materias(caminho: str) -> list[tuple[str, str]]:
    """Lê a aba Materias da planilha → [(nome_materia, serie)].
    Se a aba não existir, retorna a lista padrão hardcoded.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if "Materias" not in wb.sheetnames:
        wb.close()
        return MATERIAS_PADRAO
    ws = wb["Materias"]
    materias = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome, serie = row[0], row[1]
        if nome and serie:
            materias.append((str(nome).strip(), str(serie).strip()))
    wb.close()
    return materias if materias else MATERIAS_PADRAO