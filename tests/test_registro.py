"""Testes do registro de execução: acúmulo de linhas, truncamento de erro e export xlsx (ver ADR-003)."""

from openpyxl import load_workbook

from src.registro import COLUNAS_LOG, RegistroExecucao, resumir_erro


def test_registrar_acumula_linhas_com_timestamp():
    reg = RegistroExecucao()
    reg.registrar(login_professor="mirela.godoy", persona_id="123", fase="perfil",
                  acao="turma_adicionada", detalhe="7° EF - Turma G - Matutino", status="sucesso")
    assert len(reg.linhas) == 1
    linha = reg.linhas[0]
    assert linha["login_professor"] == "mirela.godoy"
    assert linha["acao"] == "turma_adicionada"
    assert linha["timestamp"]  # preenchido automaticamente


def test_registrar_erro_trunca_mensagem():
    reg = RegistroExecucao()
    reg.registrar_erro(ValueError("x " * 400), login_professor="a.prof", fase="validacao")
    linha = reg.linhas[0]
    assert linha["status"] == "erro"
    assert linha["acao"] == "erro"
    assert len(linha["detalhe"]) <= 300


def test_resumir_erro_achata_espacos():
    msg = resumir_erro(RuntimeError("linha1\n   linha2\t\tlinha3"))
    assert msg == "linha1 linha2 linha3"


def test_exportar_gera_xlsx_com_cabecalho_e_linhas(tmp_path):
    reg = RegistroExecucao()
    reg.registrar(login_professor="a.prof", acao="ja_existia", status="ja_configurado")
    reg.registrar(login_professor="b.prof", acao="materia_adicionada", detalhe="Matemática",
                  status="sucesso")

    caminho = reg.exportar(tmp_path)
    assert caminho.exists()
    assert caminho.parent == tmp_path
    assert caminho.name.startswith("atribuicao_") and caminho.suffix == ".xlsx"

    planilha = load_workbook(caminho).active
    valores = list(planilha.iter_rows(values_only=True))
    assert list(valores[0]) == COLUNAS_LOG       # cabeçalho
    assert len(valores) == 3                     # cabeçalho + 2 linhas
    assert valores[2][1] == "b.prof"             # coluna login_professor da 2ª linha
    assert valores[2][4] == "materia_adicionada"  # coluna acao
